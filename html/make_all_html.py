import bisect
import subprocess

import os
import sys
from collections import defaultdict
import html
import csv
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

script_dir = os.path.dirname(__file__) #<-- absolute dir the script is in
rel_path = "html_requirement.txt"
abs_file_path = os.path.join(script_dir, rel_path)

html_requirement = []
with open(abs_file_path,"r") as f:
    for line in f.readlines():
        html_requirement.append(line.replace("\n",""))

# for l in html_requirement:
#     print(l)

rel_path = sys.argv[2]
result_line_path = os.path.join(script_dir, rel_path)

rel_path = sys.argv[1]
baseline_line_path = os.path.join(script_dir, rel_path)

dir_path = result_line_path
cnt = 0
for a in reversed(result_line_path):
    if a == '/':
        break
    cnt += 1
        
dir_path  = result_line_path[:-cnt]
print(dir_path)

result_covered = defaultdict(list)
with open(result_line_path,"r") as f:
    for line in f.readlines():
        # print(line)
        idx = line.find(":")
        file_name = line[:idx]
        # print(file_name + " " + str(file_name in html_requirement))
        # print(idx)
        # file_idx = l.fined(':')
        line_num = int(line[idx+1:])
        result_covered[file_name].append(line_num)
        # print(file_name + str(line_num))

os.makedirs(dir_path+"html", exist_ok=True)
os.makedirs(dir_path+"csv", exist_ok=True)
baseline_covered = defaultdict(list)
with open(baseline_line_path,"r") as f:
    for line in f.readlines():
        # print(line)
        idx = line.find(":")
        file_name = line[:idx]
        # print(idx)
        # file_idx = l.fined(':')
        line_num = int(line[idx+1:])
        if line_num not in result_covered[file_name]:
            baseline_covered[file_name].append(line_num)
            # print(file_name + str(line_num))
        idx = line.rfind("/")
        # print(l[:idx])
        if line[:idx][-1] != '.':
            os.makedirs(dir_path+"html/"+line[:idx], exist_ok=True)
            os.makedirs(dir_path+"csv/"+line[:idx], exist_ok=True)
need = ["nested.c","vmx.c","mmu.c","x86.c"]
needfile =[]
for file_name in result_covered.keys():
    fname = file_name.split('/')
    if(fname[-1] in need):
        cov = len(result_covered[file_name])
        miss = len(baseline_covered[file_name])
        needfile.append([file_name,str(round(cov/(cov+miss)*100,1))+"%"])


head1 = '''<doctype html>
<html lang="ja">
<head>'''
head2 = '''<meta charset="utf-8">
<script src="https://cdn.rawgit.com/google/code-prettify/master/loader/run_prettify.js"></script>
<style>
    code{
        font-family:Monaco, Menlo, Consolas, 'Courier New', Courier, monospace, sans-serif;
        font-size: 14px;
        line-height: 18px;
        overflow: auto;
        resize: horizontal;
    }
    code_line{
        font-family: Monaco, Menlo, Consolas, 'Courier New', Courier, monospace, sans-serif;;
        font-size: 14px;
        line-height: 18px;
        overflow: auto;
        resize: horizontal;
        color:#303134;
    }
    blue{
        background-color:#BEEDE8;
    }
    yellow{
        background-color:#FFFF99;
    }
    red{
        background-color:#FF99AC;
    }
      .split {
         height: 100%;
         position: fixed;
         z-index: 1;
         top: 0;
         overflow-x: hidden;
      }

      .tree {
         left: 0;
         width: 20%;
      }

      .right {
         border-left: 2px solid #444;
         right: 0;
         width: 80%;
         /* font-family: 'Courier New', Courier, monospace;
				color: rgb(80, 80, 80); */
      }
</style>

</head>
<body>
   <div class="split tree">
      <ul id="file_list">
      </ul>
   </div>
   <div class="split right">
<table summary='blob content' class='blob' cellspacing="15">
<tr><td align="right"><pre><code_line>'''

end = '''
</code></pre></td></tr></table>
</div>'''


blue = PatternFill(patternType='solid', fgColor='BEEDE8')
yellow = PatternFill(patternType='solid', fgColor='FFFF99')
red = PatternFill(patternType='solid', fgColor='FF99AC')
font = Font(size = 12,name='Consolas')
right_alignment = Alignment(horizontal='right',vertical='center')
# nestedc_list = []
payload = "<script>"
payload+= "const fileList = document.getElementById('file_list')\n"
for i in range(len(needfile)):
    print(needfile[i][0],needfile[i][1])
    fname = needfile[i][0].split('/')
    payload+="fileList.innerHTML+=`<li><a href=\"/kvm_coverage/coverage/"+needfile[i][0]+".html"+"\">"+fname[-1]+" "+needfile[i][1]+"</li>`\n"
payload+="</script>"

for file_name in html_requirement:
    # csv_list = []
    with open("/home/ishii/nestedFuzz/"+file_name, "r") as f:
        file_code = f.read().split('\n')
    f = open(dir_path+"html/"+file_name+".html","w")
    # fname = file_name.split('/')
    # if(fname[-1] in need):
    #     needfile.append(file_name)
    wb = xl.Workbook()
    ws = wb['Sheet']
    # f_csv = open(dir_path+"csv/"+file_name+".csv","w")
    # csv_list.append(["line num"])
    # writer = csv.writer(f)
    # print(nested_c)
    f.write(head1)
    f.write("<title>"+file_name.split('/')[-1]+"</title>")
    f.write(head2)
    line_count = 0
    for line in file_code:
        line_count+=1
    # for i in range(1,line_count+1):
    #     f.write(str(i)+".<br>")
    f.write('''<script>for (let i = 1; i <= ''')
    f.write(str(line_count))
    f.write('''; i++){
         document.write(i+".\\n");
   }
         </script>''')
    f.write('''</code_line></pre></td>\n''')
    f.write('''<td class='lines'><pre><code class="prettyprint">''')


    n = 1
    for line in file_code:
        ws.cell(n,1,value = str(n))
        ws.cell(n,1).font = font
        ws.cell(n,2).font = font
        ws.cell(n,1).alignment = right_alignment
        ws.cell(n,2,value = "’"+line)
        if n in result_covered[file_name]:
            f.write('<blue>'+line+"</blue>"+'\n')
            ws.cell(n,2).fill = blue
        else:
            if n in baseline_covered[file_name]:
                f.write('<yellow>'+line+"</yellow>"+'\n')
                ws.cell(n,2).fill = yellow
            else:
                f.write(html.escape(line)+'\n')
        ws.row_dimensions[n].height = 18
        n+=1
    ws.column_dimensions['B'].width = 120
    ws.column_dimensions['C'].width = 60
    wb.save(dir_path+"csv/"+file_name+".xlsx")
    f.write(end)
    f.write(payload)
    f.write("</body></html>")
    f.close()